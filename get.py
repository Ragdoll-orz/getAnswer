from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
import time
import openpyxl
import os

"""
写在前面：
    Excel文件必须存在
    运行程序时不要打开Excel文件
    Excel数据不能超过2069行
    openpyxl的max_row不能获得真实的最大行数

基本实现获取题库

"""


class HashTable:
    def __init__(self, size=2069):
        self.size = size
        self.table = [None] * self.size

    def _hash(self, key):
        return hash(key) % self.size

    def set(self, key, value):
        index = self._hash(key)
        if self.table[index] is None:
            self.table[index] = [{key: value}]
        else:
            found = False
            for entry in self.table[index]:
                if key in entry:
                    entry[key] = value
                    found = True
                    break
            if not found:
                self.table[index].append({key: value})

    def get(self, key):
        index = self._hash(key)
        if self.table[index] is not None:
            for entry in self.table[index]:
                if key in entry:
                    return entry[key]
        return None


class GetAnswer:
    xlsx = r'answer.xlsx'  # 也可以绝对路径
    url = 'http://'  # 登录网址
    driver_path = r'msedgedriver.exe'  # 驱动
    # 调用WebDriver 对象的get方法 可以让浏览器打开指定网址
    driver = webdriver.Edge(executable_path=driver_path)

    def __init__(self, name, password, stage, checkboxs):
        self.name = name
        self.password = password
        self.maxRow = 0
        self.maxtime = None
        self.data = None
        self.table = None
        self.hash_table = None
        self.number = None
        # self.check_file()
        self.insert_data()
        self.login()
        self.create_question(stage, checkboxs)
        self.question_first_begin()
        self.loop_first()
        self.look_answer()
        self.question_second_begin()
        self.write_row()
        self.end()

    def check_file(self):
        if os.path.exists(self.xlsx):
            os.remove(self.xlsx)
            with open(self.xlsx, mode='a', encoding="utf-8"):
                pass
        else:
            with open(self.xlsx, mode='a', encoding="utf-8"):
                pass

    def insert_data(self):
        # 读取xlsx文件
        self.data = openpyxl.load_workbook(self.xlsx)
        # 获得指定名称页
        self.table = self.data.get_sheet_by_name('Sheet1')
        # 确定实际行数
        for row in self.table.iter_rows(min_row=1, max_row=self.table.max_row, values_only=True):
            # 如果当前行的所有单元格都为空，则认为该行为空行
            if all(cell is None for cell in row):
                continue  # 跳过空行
            else:
                self.maxRow += 1
            print("获取实际行数 ", self.maxRow)
        # 将数据塞入哈希表
        self.hash_table = HashTable()
        if self.table.cell(2, 1).value is not None:
            for a in range(2, self.maxRow + 1):
                self.hash_table.set(self.table.cell(a, 1).value, self.table.cell(a, 4).value)
                print("正向哈希表塞入数据 ", a, "/", self.maxRow)

    def login(self):
        # 窗口最大
        self.driver.maximize_window()
        #
        self.driver.implicitly_wait(6)
        # 打开网站
        self.driver.get(self.url)
        # 输入账号
        (self.driver.find_element(By.XPATH, '/html/body/div/div/div/div[2]/div/div/div/form/div[1]/div/div/div/input')
         .send_keys(self.name))
        # 输入密码
        (self.driver.find_element(By.XPATH, '/html/body/div/div/div/div[2]/div/div/div/form/div[2]/div/div/div/input')
         .send_keys(self.password))

    # 点击函数
    def click(self, xpath):
        try:
            self.driver.find_element(By.XPATH, xpath).click()
        except NoSuchElementException as e:
            print("点击函数元素未找到:", e)
        except Exception as e:
            print("点击函数发生异常:", e)
        # 1秒
        time.sleep(1)

    def create_question(self, stage, checkboxs):
        # 登录
        self.click('/html/body/div/div/div/div[2]/div/div/div/form/div[3]/div/button')
        # 考试
        self.click('/html/body/div/div/section/aside/div/div[2]/ul/li[2]/div')
        # 生成试卷
        self.click('/html/body/div[1]/div/section/aside/div/div[2]/ul/li[2]/ul/templae[2]/li')
        # 阶段
        self.click('/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[4]/div/div/div/div/div/input')
        """
        #                           阶段内容选择
        """
        # self.click('/html/body/div[2]/div[2]/div/div/div[1]/ul/li[4]')
        self.click(f'/html/body/div[2]/div[2]/div/div/div[1]/ul/li[{stage}]')
        """
        #                           levels   checkbox
        """
        for checkbox in checkboxs:
            self.click(f'/html/body/div[1]/div/section/section/main/div/main/div[2]/div[3]/div[2]/div/label[{checkbox}]')
        # self.click('/html/body/div[1]/div/section/section/main/div/main/div[2]/div[3]/div[2]/div/label[5]')
        # self.click('/html/body/div[1]/div/section/section/main/div/main/div[2]/div[3]/div[2]/div/label[6]')
        # self.click('/html/body/div[1]/div/section/section/main/div/main/div[2]/div[3]/div[2]/div/label[7]')
        # self.click('/html/body/div[1]/div/section/section/main/div/main/div[2]/div[3]/div[2]/div/label[8]')
        """
        #                           考题数量
        """
        self.click('/html/body/div[1]/div/section/section/main/div/main/div[2]/div[5]/div[2]/div/div/div/div')
        self.click('/html/body/div[2]/div[3]/div/div/div[1]/ul/li[3]')  # 100道
        # 开始考试
        self.click('/html/body/div[1]/div/section/section/main/div/main/div[2]/div[7]/button')
        # 确定考试
        self.click('/html/body/div[3]/div/div/div[3]/button[2]')
        time.sleep(1)

    # 获取文本函数
    def text(self, xpath):
        try:
            return self.driver.find_element(By.XPATH, xpath).text
        except NoSuchElementException as e:
            print("元素未找到:", e)
            return ""
        except Exception as e:
            print("发生异常:", e)

    def question_first_begin(self):
        # 获取循环次数
        str1 = self.text('/html/body/div[1]/div/section/section/main/div/main/div[2]/div[1]/ul/li[3]')
        # 最大次数
        self.maxtime = str1.rsplit(' ', 2)[2]
        # 强转
        self.maxtime = int(self.maxtime)
        self.maxtime += 1

    # 抵消js弹窗影响

    def js_alert(self, num):
        try:
            # 尝试切换到alert弹窗
            alert = self.driver.switch_to.alert
            # 如果成功切换到了alert，处理你的逻辑
            print("检测到弹窗", num)
            # 处理弹窗
            alert.accept()  # 或者可以使用 alert.dismiss() 关闭弹窗
        except:
            pass

    def loop_first(self):
        flag = True
        for a in range(1, self.maxtime):
            self.js_alert(a)

            strName = self.text('/html/body/div[1]/div/section/section/main/div/main/div[2]/div[1]/ul/li[2]')
            print(strName, a)

            if (strName == '多选题'):
                self.click('/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[2]/div/div[1]')  # 多选选A
            else:
                self.click('/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[2]/div/label[1]')  # 单选判断选A

            if flag:
                self.click('/html/body/div[1]/div/section/section/main/div/main/div[2]/div[3]/div[1]/button')  # 下一题,只有一次
                flag = False
                continue
            self.click('/html/body/div[1]/div/section/section/main/div/main/div[2]/div[3]/div[1]/button[2]')  # 下一题,最后一次消失
        self.click('/html/body/div[1]/div/section/section/main/div/div[1]/div/div/footer/div/button[1]')
        time.sleep(2)

    def look_answer(self):
        # 查看第一个解析
        self.click('/html/body/div[1]/div/section/section/main/div/main/div[2]'
              '/div/div[1]/div[3]/div/div[1]/div/table/tbody/tr[1]/td[5]/div/button[2]')

    def question_second_begin(self):
        # 获取循环次数
        str1 = self.text('/html/body/div[1]/div/section/section/main/div/main/div[2]/div[1]/ul/li[3]')
        # 最大次数
        self.maxtime = str1.rsplit(' ', 2)[2]
        # 强转
        self.maxtime = int(self.maxtime)
        self.maxtime += 1

    # 写入excel函数
    def excel(self, num, types, question, correct_answer_a, your_choice, letter, feature="", correct_answer_b="",
              correct_answer_c="", correct_answer_d=""):
        # 行，列，值 这里是从1开始计数的
        self.table.cell(num, 1, types)
        self.table.cell(num, 2, question)
        self.table.cell(num, 3, correct_answer_a)
        self.table.cell(num, 4, correct_answer_b)
        self.table.cell(num, 5, correct_answer_c)
        self.table.cell(num, 6, correct_answer_d)
        self.table.cell(num, 7, your_choice)
        self.table.cell(num, 8, letter)
        self.table.cell(num, 9, feature)

    def write_row(self):
        """
            # 写入首行
        """
        if self.table.cell(1, 1).value is None:
            self.excel(1, '题型', '题目', '正确答案', '你的错误答案', '字母', '特征')
            self.number = 2  # 让number从2开始
            self.maxRow = self.maxtime + 1
        else:
            self.number = self.maxRow + 1  # 让number从文件末尾开始
            self.maxRow = self.maxtime + self.maxRow

    # 判断题型是否存在
    def is_exist(self, str_question, str_type, str_num):
        if self.hash_table.get(str_question) is not None:
            print(str_num, " ", str_type, " 重复  ", str_question)
            return True
        else:
            return False

    @staticmethod
    def cut_str(s):
        cut_temp = int(str(len(s) / 2)[0])
        return s[cut_temp]

    def add_feature(self, aa, bb, cc, dd):
        add_a = ""
        add_b = ""
        add_c = ""
        add_d = ""
        if aa is not None:
            add_a = self.cut_str(aa)
        if bb is not None:
            add_b = self.cut_str(bb)
        if cc is not None:
            add_c = self.cut_str(cc)
        if dd is not None:
            add_d = self.cut_str(dd)
        return add_a + add_b + add_c + add_d

    def end(self):
        """
            记录答案
        """
        flag = True
        numCount = self.number  # 记录写入的位置
        for a in range(self.number, self.maxRow):
            self.js_alert(a - 1)  # 点击js弹窗
            w_types = self.text('/html/body/div[1]/div/section/section/main/div/main/div[2]/div[1]/ul/li[2]')  # 题型
            w_question = self.text('/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[1]')  # 题目
            if w_types == '判断题':
                if self.is_exist(w_question, w_types, a - 1):
                    print(a - 1, " ", w_types, " 重复  ", w_question, "~~~~~~~~~~~~~~~~~~ ", numCount)
                else:
                    w_wrong = self.text('/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[3]/span')
                    w_wrong = w_wrong.rsplit('（', 1)[1].rsplit('）', 1)[0]  # 正确or错误
                    ischeck = self.driver.find_element(By.XPATH,
                                                  '/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[2]/div/label[1]/span[1]/input').is_selected()
                    if w_wrong == "正确":
                        if ischeck:
                            self.excel(numCount, w_types, w_question, '正确', '', 'A')
                            self.hash_table.set(w_question, 'A')
                        else:
                            self.excel(numCount, w_types, w_question, '错误', '', 'B')
                            self.hash_table.set(w_question, 'B')
                        numCount += 1
                    elif w_wrong == "错误":
                        if ischeck:
                            self.excel(numCount, w_types, w_question, '错误', '正确', 'B')
                            self.hash_table.set(w_question, 'B')
                        else:
                            self.excel(numCount, w_types, w_question, '正确', '错误', 'A')
                            self.hash_table.set(w_question, 'A')
                        numCount += 1
                    else:
                        print(a - 1, "判断题判断正误出错!!!!!!!!!!!!!!!!!!")
            elif w_types == '单选题':
                if self.is_exist(w_question, w_types, a - 1):
                    w_answers_a = self.text(
                        '/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[2]/div/label[1]/span[2]')  # 答案A
                    w_answers_b = self.text(
                        '/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[2]/div/label[2]/span[2]')
                    w_answers_c = self.text(
                        '/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[2]/div/label[3]/span[2]')
                    w_answers_d = self.text(
                        '/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[2]/div/label[4]/span[2]')
                    w_feature = self.add_feature(w_answers_a, w_answers_b, w_answers_c, w_answers_d)
                    if self.is_exist(w_question + w_feature, w_types, a - 1):
                        print(a - 1, " ", w_types, " 重复*2  ", w_question, "~~~~~~~~~~~~~~~~~~ ", numCount)
                    else:
                        w_correct = self.text(
                            '/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[3]/ul[1]/li[2]')  # 正确答案
                        w_letter = w_correct[0]  # 正确答案字母
                        w_wrong = self.text(
                            '/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[3]/ul[2]/li[2]/span')  # 错误答案
                        if w_correct[0] == w_wrong:  # 不记录正确答案
                            w_wrong = ""
                        self.hash_table.set(w_question + w_feature, w_letter)
                        self.excel(numCount, w_types, w_question, w_correct, w_wrong, w_letter, w_feature)
                        print(a - 1, " ", w_question + w_feature, " ", numCount)
                        numCount += 1
                else:
                    w_correct = self.text(
                        '/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[3]/ul[1]/li[2]')  # 正确答案
                    w_letter = w_correct[0]  # 正确答案字母
                    w_wrong = self.text(
                        '/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[3]/ul[2]/li[2]/span')  # 错误答案
                    if w_correct[0] == w_wrong:  # 不记录正确答案
                        w_wrong = ""
                    self.hash_table.set(w_question, w_letter)
                    self.excel(numCount, w_types, w_question, w_correct, w_wrong, w_letter)
                    print(a - 1, " ", w_question, " ", numCount)
                    numCount += 1
            elif w_types == '多选题':
                """
                待完善错误答案记录
                """
                if self.is_exist(w_question, w_types, a - 1):
                    w_answers_a = self.text(
                        '/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[2]/div/div[1]/label/span[2]')  # 答案A
                    w_answers_b = self.text(
                        '/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[2]/div/div[2]/label/span[2]')
                    w_answers_c = self.text(
                        '/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[2]/div/div[3]/label/span[2]')
                    w_answers_d = self.text(
                        '/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[2]/div/div[4]/label/span[2]')
                    w_feature = self.add_feature(w_answers_a, w_answers_b, w_answers_c, w_answers_d)
                    if self.is_exist(w_question + w_feature, w_types, a - 1):
                        print(a - 1, " ", w_types, " 重复  ", w_question, "~~~~~~~~~~~~~~~~~~ ", numCount)
                    else:
                        # /html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[3]/ul[1]/li[2]/span[1]
                        # /html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[3]/ul[1]/li[3]/span[1]
                        w_letter = ""
                        for b in range(2, 6):
                            temp = self.text(
                                '/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[3]/ul[1]/li[' + str(
                                    b) + ']/span[1]')
                            if temp != "":
                                w_letter += temp
                            else:
                                z = 1
                        w_correct_a = self.text(
                            '/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[3]/ul[1]/li[2]')  # 正确答案
                        w_correct_b = self.text(
                            '/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[3]/ul[1]/li[3]')  # 正确答案
                        w_correct_c = self.text(
                            '/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[3]/ul[1]/li[4]')  # 正确答案
                        w_correct_d = self.text(
                            '/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[3]/ul[1]/li[5]')  # 正确答案
                        w_wrong = self.text(
                            '/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[3]/ul[2]/li[2]/span')  # 错误答案
                        self.hash_table.set(w_question + w_feature, w_letter)
                        self.excel(numCount, w_types, w_question + w_feature, w_correct_a, w_wrong, w_letter, w_feature,
                              w_correct_b,
                              w_correct_c, w_correct_d)
                        print(a - 1, " ", w_question + w_feature, " ", numCount)
                        numCount += 1
                else:
                    w_letter = ""
                    for b in range(2, 6):
                        temp = self.text(
                            '/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[3]/ul[1]/li[' + str(
                                b) + ']/span[1]')
                        if temp != "":
                            w_letter += temp
                        else:
                            z = 1
                    w_correct_a = self.text(
                        '/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[3]/ul[1]/li[2]')  # 正确答案
                    w_correct_b = self.text(
                        '/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[3]/ul[1]/li[3]')  # 正确答案
                    w_correct_c = self.text(
                        '/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[3]/ul[1]/li[4]')  # 正确答案
                    w_correct_d = self.text(
                        '/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[3]/ul[1]/li[5]')  # 正确答案
                    w_wrong = self.text(
                        '/html/body/div[1]/div/section/section/main/div/main/div[2]/div[2]/div[3]/ul[2]/li[2]/span')  # 错误答案
                    self.hash_table.set(w_question, w_letter)
                    self.excel(numCount, w_types, w_question, w_correct_a, w_wrong, w_letter, "", w_correct_b, w_correct_c,
                          w_correct_d)
                    print(a - 1, " ", w_question, " ", numCount)
                    numCount += 1
            else:
                print(a - 1, "题 题型判断出错!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")

            if flag:
                self.click('/html/body/div[1]/div/section/section/main/div/main/div[2]/div[1]/div/button')  # 下一题,只有一次
                flag = False
                continue
            if a != self.maxRow - 1:
                self.click('/html/body/div[1]/div/section/section/main/div/main/div[2]/div[1]/div/button[2]')  # 下一题,最后一次消失
        time.sleep(2)  # 防止末尾的数据写不进去
        self.data.save(self.xlsx)  # 保存excel
        self.driver.quit()  # 退出浏览器


# if __name__ == '__main__':
#     ga = GetAnswer('账号', '密码', 2, [1, 2])
